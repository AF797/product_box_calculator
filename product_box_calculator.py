import tkinter as tk
import openpyxl

def search_excel():
    product_name = entry_product_name.get()
    search_keywords = product_name.split()  # 공백을 기준으로 검색어를 분리

    # 엑셀 파일 열기
    wb = openpyxl.load_workbook('example.xlsx')

    # 검색 결과 저장할 리스트
    results = []

    # 모든 시트 탐색
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # 엑셀에서 정보 검색 (B열만 탐색)
        for row in sheet.iter_rows(values_only=True):
            cell_value = str(row[1]).replace(" ", "")  # B열 값 가져오기

            if all(keyword in cell_value for keyword in search_keywords):
                results.append((row[1], row[6]))  # B열(row[1])과 G열(row[6]) 값 저장

    # 결과 출력
    if results:
        result_text = "<결과>\n"
        for row in results:
            result_text += str(row[0]) + "EA, " + str(row[1]) + "EA" + '\n'  # B열(row[0])과 G열(row[1]) 값을 출력
        label_result.config(text=result_text)
    else:
        label_result.config(text="일치하는 제품을 찾을 수 없습니다.")

    # 엑셀 파일 닫기
    wb.close()


# 창 생성
window = tk.Tk()
window.title("규림 제품 갯수 계산기")

# 아이콘 파일 경로
icon_path = "kr.ico"

# 아이콘 설정
window.iconbitmap(default=icon_path)

# 창의 크기 설정
window.geometry("300x500")

# 안내 문구를 표시할 라벨 생성
label = tk.Label(window, text="제품 갯수 계산기")
label.place(relx=0.5, rely=0.1, anchor=tk.CENTER)  # 가운데 정렬

label = tk.Label(window, text="입력 예시 : 제리캔 5L, 제리캔5L")
label.place(relx=0.5, rely=0.2, anchor=tk.CENTER)

label = tk.Label(window, text="made by Min")
label.place(relx=0.15, rely=0.02, anchor=tk.CENTER)

# 입력 상자 생성
entry_product_name = tk.Entry(window)
entry_product_name.place(relx=0.5, rely=0.35, anchor=tk.CENTER)  # 가운데 정렬

# 버튼 생성
button = tk.Button(window, text="검색", command=search_excel)
button.place(relx=0.5, rely=0.4, anchor=tk.CENTER)  # 가운데 정렬

# 결과 출력 라벨 생성
label_result = tk.Label(window, text="")
label_result.place(relx=0.5, rely=0.7, anchor=tk.CENTER)


# 이벤트 루프 시작
window.mainloop()
